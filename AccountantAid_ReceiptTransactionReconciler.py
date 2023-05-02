#!/usr/local/bin/python3.10
import pandas as pd # import pandas module to convert csv to xlsx
from openpyxl import load_workbook # import openpyxl to parse and modify csv
import os

def fileNameEditor(filename, card, vendorDict):
    
    itemsToExclude = [".pdf", ".jpg", ".jpeg", ".z", "-", "---", ".png"]
    for item in itemsToExclude:
        filename = filename.replace(item, "")
    filename = filename.lstrip()
    filename = filename.rstrip()
    fileNameArray = filename.split(" ")

    indicators = ["verified", "DiffByDateOnly", "DiffNOTOnlyByDate"]
    for verStat in indicators:
        if verStat in filename:
            verStat = verStat
        else:
            verStat = ""

    # checks for venderShorthand in file name at position 1
    try:
        vendorShorthand = fileNameArray[1]
    except:
        IndexError

    # checks if selected Card is in fileName for reconciling
    if card in fileNameArray:
        if vendorShorthand not in list(vendorDict.keys()):
            for key in tuple(vendorDict.keys()):
                if key not in vendorShorthand:
                    continue
                else:
                    vendorShorthand =  key

        # cleans up file
        date = fileNameArray[0]
        date = fileNameArray[0].replace(":", "-")
        date = fileNameArray[0].replace("_", "-")

        # remves preceedign dash r underscores by assigning position first int
        for i in range(len(date)):
            if date[i] != "-" or date[i] != "_":
                pos = i
                break
        date = date[pos:]
        
        # checks for single digit months wo 0 & adds 0 (4-11-23 to 04-11-23)
        if date[1] == "-":
            date = "0" + date

        # checks for single digit days wo 0 & adds 0 (12-1-23 to 12-01-23)
        if date[2] == "-" and date[4] == "-":
            date = date[:3] + "0" + date[3:]

        # moves year to front
        date = date[:4] + date[4:].replace("-22", "-2022")
        date = date.replace("-2022", "")
        date = "2022-" + date
    
        # removes ending 0s
        amount = fileNameArray[2].replace("$", "")
        if ".00" in amount:
            amount = amount[0:-3]
        if "." in amount and amount[-1] == "0":
            amount = amount[0:-1]
            if "." in amount and amount[-1] == "0":
                amount = amount[0:-1]
                if "." in amount and amount[-1] == ".":
                    amount = amount[0:-1]

        if False:
            print("Files initial:", date, vendorShorthand, amount, card, verStat)
                  
        return(date, vendorShorthand, amount, card, verStat)


class Compare():
    def __init__(self):
        self.excelFile = ""
        self.receipts = ""
        self.transactions = ""
        self.card = ""
        self.vendorDict = {}
        self.receiptFolder = ""
        self.matched = set()
        self.set1Transactions = set()
        self.matchednewCatCoord = []
        self.unmatchednewCatCoord = set() # holds currentCategory & newCatCoord 
        self.unmatchedDiffOnlyByDateCoord = set() # holds unmatched DiffOnlyByDate
        self.unmatchedTransByDateOnly_VndAmtSet = set() # transaction vendor and amount to check if unmatched is by date only
        self.FileNotMatchedWTrans = set()

    def __str__(self):
        return str(self.receipts + '\t' + self.transactions)
    
    # stores receipts
    def storeReceipts(self, receipts, receiptFolder):
        self.receipts = receipts
        self.receiptFolder = receiptFolder
        return self.receipts
    
    # stores transactions and last row, which gives range to fill out empty fields automatically
    def storeTransactions(self, transactions, lastRow, excelfile):
        self.excelFile = excelfile
        self.transactions = transactions
        self.lastRow = lastRow
        return self.transactions

    def storeCard(self, card):
        self.card = card
        return self.card
    
    def storeVendors(self, vendorDict):
        self.vendorDict = vendorDict
        return self.vendorDict
    
    def storeBusinessName(self, BusinessName):
        self.BusinessName = BusinessName
        return self.BusinessName
    
    def storefolderExcel(self, folderExcel):
        self.folderExcel = folderExcel
        return self.folderExcel 

    def process(self):
        set1TransDtVndAmt = set() # transaction Date, Vendor, Amount
        receiptDtVndAmt = set() # locally used set for receipts

        # Transactions
        for item in self.transactions:
            # print("self.transactions", item[0], item[1], item[2], item[4], item[5], item[6])
            #postDate, vendorShorthand, amount, card, Category, Include, newCategory
            self.set1Transactions.add((item[0], item[1], item[2], item[3], item[4], item[5], item[6]))

            #postDate, vendorShorthand, amount
            set1TransDtVndAmt.add((item[0], item[1], item[2]))

            # transaction vendor and amount to check if unmatched is by date only
            self.unmatchedTransByDateOnly_VndAmtSet.add((item[1], item[2]))
        
        # Receipt fields: date, vendorShorthand, amount, card, ver
        for receiptItem in self.receipts:
            # print(receiptItem)
            receiptDtVndAmt.add((receiptItem[0], receiptItem[1], receiptItem[2]))

        if (False):
            for transactions in self.transactions:
                print("self.transactions", transactions)
            else:
                print("self.transactions is blank.")

        if (False):
            for set1items in set1TransDtVndAmt:
                print("set1TransDtVndAmt", set1items)
            else:
                print("No transactiosn received!")

        if (False):
            for set2items in receiptDtVndAmt:
                print("receiptDtVndAmt", set2items)



        self.matched = set1TransDtVndAmt & receiptDtVndAmt
        unmatched1 = set1TransDtVndAmt.difference(receiptDtVndAmt) # transactions exist but not files
        unmatched2 = receiptDtVndAmt.difference(set1TransDtVndAmt) # have file but not transaction

        if len(self.matched) == 0:
            print("ERROR: NO RECEIPTS MATCHED. CHECK RECEIPT FOLDER: If runnning program for 2nd time, be sure receipts are in first directy of ReceiptFolder. They may have been placed in 'verified' or another folder.")
            return

        for item in self.set1Transactions:
            # get the coordinate for newCategory row of cell that matched
            if (item[0], item[1], item[2]) in self.matched:
                #print("matched", item[0], item[1], item[2])
                cellCoord = str(item[6]).split(".")[1].replace(">", "")
                self.matchednewCatCoord.append(cellCoord)
            else: # same as unmatched 1
            # ones that don't match are put in unmatched.
                #print("TransactionsUnmatched", item[0], item[1], item[2], item[4], item[5])
                cellCoordCurrCat = str(item[4]).split(".")[1].replace(">", "")
                cellCoordInc = str(item[5]).split(".")[1].replace(">", "")
                cellCoordNewCat = str(item[6]).split(".")[1].replace(">", "")

                self.unmatchednewCatCoord.add((cellCoordCurrCat, cellCoordInc, cellCoordNewCat))
                #print(cellCoord)

        # have file but not transaction
        for item in unmatched2:
            if (item[1], item[2]) in self.unmatchedTransByDateOnly_VndAmtSet:
                print("FileNotMatchedWTrans DiffOnlyByDate")
                print("\tfile:", item)
                for transaction in self.set1Transactions:
                    # print(transaction)
                    for values in transaction:
                        if item[2] in str(values):
                            print("\ttrns:", transaction)
                            cellCoordCurrCat = str(transaction[4]).split(".")[1].replace(">", "")
                            cellCoordInc = str(transaction[5]).split(".")[1].replace(">", "")
                            cellCoordNewCat = str(transaction[6]).split(".")[1].replace(">", "")
                            self.unmatchedDiffOnlyByDateCoord.add((cellCoordCurrCat, cellCoordInc, cellCoordNewCat))

        # have file but not transaction
        # separated from above for easy viewing
        for item in unmatched2:
            if (item[1], item[2]) not in self.unmatchedTransByDateOnly_VndAmtSet:
                    print("FileNotMatchedWTrans DiffNOTOnlyByDate" , item)
                    self.FileNotMatchedWTrans.add(item)


        # displays results if True for debugging
        if (False):     
            for item in unmatched1: # transactions exist but not files
                print("unmatched1", item)

        if (False):   
            for item in unmatched2: # transactions exist but not files
                print("unmatched2", item)
            print()

        C.updateExcel()
        C.updateReceipts()

    def updateExcel(self):
        try:
            book = load_workbook(self.excelFile)
            sheet = book.active  # iterable
        except:
            print("Error: Excel Not Found. Check Path or FileName ")
            print("ExcelLocation:", self.excelFile)
            FileNotFoundError
        
        # adds matched to excel
        for coordMatched in self.matchednewCatCoord:
            sheet[f'{coordMatched}'] = f"{self.BusinessName} 12) Repair - Verified"
        
        # adds unmatched to excel
        for (cellCoordCurrCat, cellCoordInc, cellCoordNewCat) in self.unmatchednewCatCoord:
            #print(sheet[cellCoordCurrCat].value.lower())
            #print("coordUnMatched", cellCoordCurrCat, cellCoordNewCat)
            try:
                # if verified already, cellCoordNewCat, keeps that there
                if "verified".lower() in sheet[cellCoordNewCat].value.lower():
                    sheet[f'{cellCoordNewCat}'] = sheet[cellCoordNewCat].value
                else:
                    if "Exc".lower() in sheet[cellCoordInc].value.lower():
                        #print("Exc", sheet[cellCoordInc].value, cellCoordInc)
                        sheet[f'{cellCoordNewCat}'] = sheet[cellCoordInc].value
                        # keep sometimes can be both, we want to use below if both
                    if "Payment".lower() in sheet[cellCoordCurrCat].value.lower():
                        #print("Credits:", sheet[cellCoordCurrCat].value, cellCoordNewCat)
                        sheet[f'{cellCoordNewCat}'] = "NA - PaymentORCredit"
                    if "Exc".lower() not in sheet[cellCoordInc].value.lower() and "Payment".lower() not in sheet[cellCoordCurrCat].value.lower():
                        sheet[f'{cellCoordNewCat}'] = f"{self.BusinessName} 12) Repair - NoReceiptFound"
                    if "YES_IR_".lower() in sheet[cellCoordInc].value.lower():
                        sheet[f'{cellCoordNewCat}'] = sheet[cellCoordCurrCat].value
            except:
                AttributeError
        
        # when unmatched diff by date, notes in excel
        # if "exc" in cellCoordInc, assigns it to cellCoordNewCat
        for (cellCoordCurrCat, cellCoordInc, cellCoordNewCat) in self.unmatchedDiffOnlyByDateCoord:
            # if verified already, cellCoordNewCat, keeps that there
            if "verified".lower() in sheet[cellCoordNewCat].value.lower():
                sheet[f'{cellCoordNewCat}'] = sheet[cellCoordNewCat].value
            else:
                # print("self.unmatchedDiffOnlyByDateCoord", cellCoordCurrCat, cellCoordInc, cellCoordNewCat)
                if "exc" not in sheet[f'{cellCoordInc}'].value:
                    sheet[f'{cellCoordNewCat}'] = f"{self.BusinessName} 12) Repair - DiffByDateOnly"
                else:
                    sheet[f'{cellCoordNewCat}'] = sheet[f'{cellCoordInc}'].value

                # sets cellCoordNewCategory with cellCoordCurrCat if unmatched
        for rows in range(1, self.lastRow):
            try:
                if (sheet[f'{cellCoordNewCat[0]+ str(rows)}'].value) == None:
                    sheet[f'{cellCoordNewCat[0]+ str(rows)}'] = sheet[f'{cellCoordInc[0]+ str(rows)}'].value
            except:
                UnboundLocalError

        # write to file and create a new filename with "updated" appended
        # if a previously updated file is in folder appends the sum to end of "updated" when making filename
        if ".xl" in self.excelFile: # checks file has xlsx
            #print(sheet[f'{cellCoordNewCat}'].value)
            docNameNew = self.excelFile.split(".x")
            filenames = os.listdir(self.folderExcel)

            _updatedList = 0
            for item in filenames:
                if "_updated" in item:
                    _updatedList += 1

            if _updatedList != 0:
                    print(str(_updatedList+1) + " updated files found in folder will NOT be deleted/modified. ")
                    fileUpdated = docNameNew[0] + "_updated_" + str(_updatedList+1) + ".x" + docNameNew[1]
                    # print("fileUpdated", fileUpdated)
                    book.save(filename=f'{fileUpdated}')
            else:
                fileUpdated = docNameNew[0] + "_updated_0" + ".x" + docNameNew[1]
                book.save(filename=f'{fileUpdated}')

            print("\n")
            print("PROCESS COMPLETED. Check above for any errors. And open files below to confirm.")
            print("Excel location:", self.folderExcel)
            print("Receipts location: ", self.receiptFolder)
            print("\n")
            print("x"*100)
            print("\n")
        else:
            print("Please use excel file")

    def updateReceipts(self):
        fileExt = [".pdf", ".jpg", ".jpeg", ".png"]
        indicators = ["verified", "DiffByDateOnly", "DiffNOTOnlyByDate"]
        filenames = os.listdir(self.receiptFolder)
        for filename in filenames:
            receiptsSet = fileNameEditor(filename, self.card, self.vendorDict) # creates tuple per filename & returns a set()
            if indicators[0] not in filename and indicators[1] not in filename and indicators[2] not in filename:
                if receiptsSet is not None and (receiptsSet[0], receiptsSet[1], receiptsSet[2]) in self.matched:
                    # print("MatchedFile: ", fileName, " :\t ", receiptsSet[0], receiptsSet[1], receiptsSet[2])
                    for item in fileExt:
                        if item in filename:
                            startFile = self.receiptFolder + "/" + filename
                            # print("startFile", startFile)
                            try:
                                C.makeDirectory("verified")
                            except:
                                FileExistsError
                            endFile = self.receiptFolder + "/verified/" + filename.replace(item, " verified"+item)
                            # print("endFile", endFile)
                            os.rename(startFile, endFile)
                elif receiptsSet is not None and (receiptsSet[1], receiptsSet[2]) in self.unmatchedTransByDateOnly_VndAmtSet:
                    for item in fileExt:
                        if item in filename:
                            startFile = self.receiptFolder + "/" + filename
                            # print("startFile", startFile)
                            try:
                                C.makeDirectory("DiffByDateOnly")
                            except:
                                FileExistsError
                            endFile = self.receiptFolder + "/DiffByDateOnly/" + filename.replace(item, " DiffByDateOnly"+item)
                            # print("endFile", endFile)
                            os.rename(startFile, endFile)
                elif receiptsSet is not None and (receiptsSet[0], receiptsSet[1], receiptsSet[2]) in self.FileNotMatchedWTrans:
                    for item in fileExt:
                        if item in filename:
                            startFile = self.receiptFolder + "/" + filename
                            # print("startFile", startFile)
                            try:
                                C.makeDirectory("DiffNOTOnlyByDate")
                            except:
                                FileExistsError
                            endFile = self.receiptFolder + "/DiffNOTOnlyByDate/" + filename.replace(item, " DiffNOTOnlyByDate"+item)
                            # print("endFile", endFile)
                            os.rename(startFile, endFile)

    def makeDirectory(self, verStatus):
        verStatus = verStatus
        # print("receiptFolder", self.receiptFolder)
        path = os.path.join(self.receiptFolder, verStatus)
        # print("path", path)
        os.mkdir(path)

C = Compare()

class InitialScreen():

    def __init__(self):
        self.BusinessName = ""
        self.vendorDict = {}
        self.cardNames = []
        self.cardNameSelected = ""
        self.excelFile = ""
        self.receiptsFolder = ""
        self.receiptFolderSpecificToCard = ""

    def __str__(self):
        return str(self.BusinessName + '\t' + self.BusinessName)

    def setCardName(self, resp):
        self.cardNames = ["Discover5658", "Chase5726", "Chase7208", "Amex1005", "Amex1006", "Cash"]
        self.cardNameSelected = self.cardNames[resp]
        return self.cardNameSelected
    
    def setVendorDict(self):
        self.vendorDict = {
                        "HDT": ["HDT", "home depot", "HOMEDEPOT.COM", "HOME DEPOT"], 
                        "HDWE": ["HDWE", "AceHardware"], 
                        "AMZN": ["AMZNCPHT", "Amazon"],
                        "HVBCPCL": ["HVBCPCL"],
                        "TLRP": ["TLRP"],
                        "CST": ["COSTCO"],
                        "EB": ["EBAY"],
                        }
        return self.vendorDict
    
    def setReceiptFolder2(self, receiptFolderSpecificToCard):
        self.receiptFolderSpecificToCard = receiptFolderSpecificToCard
        return self.receiptFolderSpecificToCard
        
    def welcomeMessage(self):
            # prints welcome message
        print("\n")
        print("x"*100)
        print("""
        Welcome Accountant Aid: Receipts-Transaction Reconciler
        This program reconciles receipts with bank account transations & 
        adds 'verified, 'DiffOnlyByDate', 'DiffNotOnlyByDate' to transactions and receipts.

        Step 1: Select appropriate business.
        Step 2: Be sure transaction file pathname matches (remove x).
        Step 3: Reivew ERROR reasons closely & troubleshoot. 
        
        Prep: Be sure transactionsn have Date, Vendor, and Amount. Remove any spece or symbols before date
        Correct setup example: 10_21_22 HDT $153.62 2022Receipt_34_4

        Unhandled Error Note: If you have space preceeding date, it will not work "___ ". Fix for future versions.
            """
        )
        I.checkBusiness()
            
    def checkBusiness(self):

        # receives input from business selection & fixes errors
        try:
            response = int(input(str(">Select from database: 0: HMRDept 1: IRDept:  ")))
            if response == 0:
                self.BusinessName = "HMRDept"
                print("You have selected:", self.BusinessName)
                inp = input("press <enter> to continue")
                I.createFolderLocation()
            elif response == 1:
                self.BusinessName == "HMRDept"
                I.createFolderLocation()
            else:
                print("Please only enter '0' or '1'")
                I.checkBusiness()
        except:
            print("Please only enter '0' or '1'") # notifies user of error
            I.checkBusiness() # when error occurs it restarts function
            ValueError
            return
        
    def createFolderLocation(self):
        self.receiptsFolder = input("Enter folder with receipts: ")
        I.checkReceiptFolder(self.receiptsFolder)

    #  tests to make sure receiptFolder exists
    def checkReceiptFolder(self, receiptFolder):
        try:
            os.listdir(receiptFolder)
            print("Found Receipt Folder: ", receiptFolder)
            print("\t")
            I.setUpCardFilePaths()
            # return
        except:
            print("\nERROR: ReciptFolder not found. Check folder location to to make sure names are as follows and receipts are not in folders.")
            print("Provided folder:", self.receiptsFolder)
            FileNotFoundError
            return

    # sets the excel file that contains the transaction
    def setTransactionFile(self):
            excelFile = input("Enter transaction file path: ")
            input("press <enter> to continue")
            if ".xlsx" not in excelFile:
                print("ERROR: No transaction file found")
                return
            self.excelFile = excelFile
            return self.excelFile
    
    # sets up card to be used and the path of the excel and receipt folder
    def setUpCardFilePaths(self):
        I.setVendorDict() # sets dictionary of vendors
        
        # asks user to select card type
        try:
            resp = int(input(str(">Choose type: 0:Discover5658 1:Chase5726 2:Chase7208 3:Amex1005 4:Amex1006 5:Cash ")))
            if 0 > resp > 5:
                print("Error: Please only enter 0 to 5")
                I.setUpCardFilePaths()
        except:
            print("Error: Please only enter 0 or 5")
            I.setUpCardFilePaths()
            ValueError

        card = I.setCardName(resp)
        excelFile = I.setTransactionFile()

        print("You have selected:", card)
        inp = input("press any key to continue")

        # extracts folder containing transactions from excelFile path
        # this is used to check if an updated file already exists, and if so not to overwrite it
        excelFilePath = excelFile
        excelFilePath = excelFilePath.split("/")
        excelFilePath = excelFilePath[0:-1]

        folderExcel = ""
        for section in excelFilePath:
            folderExcel += "/" + section
        folderExcel = folderExcel.replace("//", "/")
        print("folderExcel", folderExcel)
        
        I.gatherTransactionFiles(excelFile, self.receiptsFolder, self.setCardName(resp), self.vendorDict, self.BusinessName, folderExcel)
    
    # iterates through excel and puts transactions in a set
    def gatherTransactionFiles(self, excelFile, receiptFolder, card, vendorDict, BusinessName, folderExcel):
        sheet = ""
        lastRow = ""
        transactionsSet = set()
        
        # checks to make sure excel exsists
        try:
            book = load_workbook(excelFile)
            sheet = book.active
        except:
            print("ERROR: Excel Not Found. Check Path or FileName ")
            print("ExcelLocation:", excelFile)
            FileNotFoundError
        
        # gets the last row
        for row in sheet:
            for cell in row:
                if cell.value is not None:
                    lastRow = cell.row
        
        vendors = [] # captures 

        for vendorShorthand, vendor in vendorDict.items():
            if isinstance(vendor, list):
                for item in vendor:
                    vendors.append(item)
 
    
        # iterates through transactions & creates a set for each row
        for row in sheet:
            for cell in row:
                try:
                    if 'Description' in cell.value:
                        start_point = "" + cell.coordinate[0] + str(cell.row+1) + ""
                        end_pointEstimate = "" + cell.coordinate[0] + str(cell.row+lastRow+1) + ""
                        for row in sheet[f'{start_point}': f'{end_pointEstimate}']:
                        # for row in sheet['D281':'D281']:
                            for cell in row:
                                if cell.value is not None:
                                    for vendor in vendors:
                                        # print("vendor", vendor.lower(), cell.value.lower())
                                        if vendor.lower() in cell.value.lower():
                                            #print("found", vendor, vendorShorthand, cell.value)
                                            foundRow = cell.row
                                            if (sheet["E1"].value) != "Amount":
                                                break
                                            if (sheet["F1"].value) != "Card":
                                                break
                                            if (sheet["G1"].value) != "Category":
                                                break
                                            if (sheet["H1"].value) != "Include":
                                                break
                                            if (sheet["J1"].value) != "NewCategory":
                                                break
                                            postDate = str(sheet["C"+f'{foundRow}'].value).split(" ")[0]
                                            amount = str(sheet["E"+f'{foundRow}'].value)
                                            amount = str(abs(float(amount))) # gets rid of negative numbers
                                            if ".00" in amount:
                                                amount = amount[0:-3]
                                            elif ".0" in amount and amount[-1] == "0":
                                                amount = amount[0:-2]
                                            elif "." in amount and amount[-1] == ".":
                                                amount = amount[0:-1]
                                            
                                            card = sheet["F"+f'{foundRow}'].value
                                            Category = sheet["G"+f'{foundRow}']
                                            Include = sheet["H"+f'{foundRow}']
                                            newCategory = sheet["J"+f'{foundRow}']

                                            # assigns vendorShorthand
                                            for key, value in vendorDict.items():
                                                if isinstance(value, list):
                                                    for item in value:
                                                        if item == vendor:
                                                            vendorShorthand = key
                                            
                                            if (False):
                                                print("Transactions", postDate, vendorShorthand, vendor, amount, card, Category, newCategory)
                                            transactionsSet.add((postDate, vendorShorthand, amount, card, Category, Include, newCategory))

                except TypeError:
                    continue

        # fetches & cleans up receipt names
        filenames = os.listdir(receiptFolder)  # assigns files in receiptFolder to filenames
        receiptsSet = set()
        for filename in filenames:  # goes through individual files
            if fileNameEditor(filename, card, vendorDict) is not None:
                receiptsSet.add(fileNameEditor(filename, card, vendorDict))

         # sends transactions for comparison
        I.organizeTransactions(transactionsSet, lastRow, excelFile)

        #stores in compare Compare()
        I.organizeFileNames(receiptsSet, receiptFolder, card, vendorDict, BusinessName, folderExcel)

        # begins comparing transactions with receipts
        I.beginCompare()

    # stores excel transactions, # of rows, and file location
    def organizeTransactions(self, transactions, lastRow, excelfile) -> object:
        C.storeTransactions(transactions, lastRow, excelfile)

    # stores receipt related values
    def organizeFileNames(self, receiptsSet, receiptFolder, card, vendorDict, BusinessName, folderExcel) -> object:
        C.storeReceipts(receiptsSet, receiptFolder)
        C.storeCard(card)
        C.storeVendors(vendorDict)
        C.storeBusinessName(BusinessName)
        C.storefolderExcel(folderExcel)

    def beginCompare(self):
        C.process() # initializes the comparison


# creates object from class that gets use input
I = InitialScreen()

# function starts program by getting business info
def initialize() -> object:
    I.welcomeMessage()

# starts prpogram
initialize()
